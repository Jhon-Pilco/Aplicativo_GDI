"""
Vista del Módulo de Reportes conectada a la BD
"""
# reports_view.py
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from styles import Colors, Fonts

# matplotlib para gráficos y exportar a PDF (tabla como figura)
import matplotlib
matplotlib.use("Agg")  # backend para guardar figuras (PDF/PNG) sin abrir ventana
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# intentar importar openpyxl (Excel); si falla, lo detectamos y deshabilitamos el botón
try:
    import openpyxl
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


class ReportsView:
    """Vista del módulo de reportes (con export Excel/PDF y gráficos embebidos)."""

    def __init__(self, parent, db_controller, back_callback=None):
        """
        parent: frame donde se incrusta la vista (HomeView pasa content_area)
        db_controller: instancia de DBController (tiene fetchall)
        back_callback: función opcional a llamar al pulsar INICIO (normalmente HomeView.show_inicio)
        """
        self.parent = parent
        self.db = db_controller
        self.back_callback = back_callback

        # Estado actual (columnas/filas) para exportar/graficar
        self.current_columns = []
        self.current_rows = []

        # Frame principal
        self.main_frame = tk.Frame(parent, bg=Colors.SURFACE)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        self.create_widgets()
        # No cargamos un reporte por defecto automáticamente: el usuario elegirá.
        # Pero si quieres que cargue el 3ro por defecto, descomenta la siguiente línea:
        # self.load_report_data()

    # -----------------------------
    # CREAR WIDGETS (INTERFAZ)
    # -----------------------------
    def create_widgets(self):
        header_frame = tk.Frame(self.main_frame, bg=Colors.SURFACE)
        header_frame.pack(fill=tk.X, pady=(0, 12))

        # Botón INICIO (usa back_callback si fue provisto)
        inicio_btn = tk.Button(
            header_frame, text="INICIO", font=Fonts.BODY,
            bg=Colors.BACKGROUND, fg=Colors.TEXT,
            relief=tk.RAISED, bd=1, padx=14, pady=6, cursor="hand2",
            command=self._handle_back
        )
        inicio_btn.pack(side=tk.LEFT)

        title = tk.Label(self.main_frame, text="Reportes de Clientes", font=Fonts.SUBTITLE,
                         bg=Colors.SURFACE, fg=Colors.TEXT)
        title.pack(pady=(0, 12))

        content_frame = tk.Frame(self.main_frame, bg=Colors.SURFACE)
        content_frame.pack(fill=tk.BOTH, expand=True)

        # ------------------ PANEL IZQUIERDO ------------------
        left_panel = tk.Frame(content_frame, bg=Colors.BACKGROUND, relief=tk.SOLID, bd=1, width=320)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 18))
        left_panel.pack_propagate(False)

        # Buscar/filtrar rápido
        search_frame = tk.Frame(left_panel, bg=Colors.BACKGROUND)
        search_frame.pack(fill=tk.X, pady=(12, 6), padx=10)

        tk.Label(search_frame, text="Buscar texto:", font=Fonts.SMALL, bg=Colors.BACKGROUND, fg=Colors.TEXT).pack(anchor=tk.W)
        self.quick_search = tk.Entry(search_frame, font=Fonts.BODY)
        self.quick_search.pack(fill=tk.X, pady=6)
        self.quick_search.bind("<Return>", lambda e: self._apply_quick_search())

        tk.Button(search_frame, text="Aplicar", font=Fonts.SMALL, command=self._apply_quick_search).pack(pady=(0,6))

        # Lista de reportes con scrollbar
        list_frame = tk.Frame(left_panel, bg=Colors.BACKGROUND)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        self.report_listbox = tk.Listbox(list_frame, font=Fonts.SMALL, activestyle='dotbox',
                                         selectmode=tk.SINGLE, exportselection=False, height=12)
        self.report_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.report_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.report_listbox.config(yscrollcommand=scrollbar.set)

        # llenar la listbox
        self.reports_list = [
            "1. Clientes nuevos por fechas festivas",
            "2. Ranking de regiones activas",
            "3. Crecimiento anual de clientes",
            "4. Preferencias de clientes minoristas",
            "5. Clientes sin contrato completo",
            "6. Cobertura de clientes corporativos y mayoristas",
            "7. Distribución de correos electrónicos",
            "8. Contratos activos por cliente corporativo",
            "9. Administradores con más clientes corporativos activos",
            "10. Ranking de administradores según clientes gestionados"
        ]
        for r in self.reports_list:
            self.report_listbox.insert(tk.END, r)
        # seleccionar por defecto el 3 (índice 2)
        self.report_listbox.select_set(2)
        self.report_listbox.bind("<<ListboxSelect>>", lambda e: self._on_list_select())

        # ------------------ PANEL DERECHO ------------------
        right_panel = tk.Frame(content_frame, bg=Colors.SURFACE)
        right_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # botones arriba
        buttons_frame = tk.Frame(right_panel, bg=Colors.SURFACE)
        buttons_frame.pack(fill=tk.X, pady=(0,8))

        gen_btn = tk.Button(buttons_frame, text="Generar Reporte", font=Fonts.BUTTON,
                            bg=Colors.PRIMARY, fg="white", command=self.load_report_data, cursor="hand2")
        gen_btn.pack(side=tk.LEFT, padx=(0,8))

        csv_btn = tk.Button(buttons_frame, text="Descargar CSV", font=Fonts.BUTTON,
                            bg=Colors.BACKGROUND, fg=Colors.TEXT, command=self.download_report, cursor="hand2")
        csv_btn.pack(side=tk.LEFT, padx=8)

        # Excel
        self.excel_btn = tk.Button(buttons_frame, text="Exportar Excel", font=Fonts.BUTTON,
                                   bg=Colors.BACKGROUND, fg=Colors.TEXT, command=self.export_to_excel, cursor="hand2")
        self.excel_btn.pack(side=tk.LEFT, padx=8)
        if not _HAS_OPENPYXL:
            # Deshabilitar y mostrar tooltip breve (mensaje al click)
            self.excel_btn.config(state=tk.DISABLED)
            self.excel_btn.bind("<Button-1>", lambda e: messagebox.showwarning(
                "openpyxl faltante",
                "Para exportar a Excel instale:\n\npip install openpyxl"
            ))

        pdf_btn = tk.Button(buttons_frame, text="Exportar PDF", font=Fonts.BUTTON,
                            bg=Colors.BACKGROUND, fg=Colors.TEXT, command=self.export_to_pdf, cursor="hand2")
        pdf_btn.pack(side=tk.LEFT, padx=8)

        graph_btn = tk.Button(buttons_frame, text="Ver Gráfico", font=Fonts.BUTTON,
                              bg=Colors.BACKGROUND, fg=Colors.TEXT, command=self.show_graph_window, cursor="hand2")
        graph_btn.pack(side=tk.LEFT, padx=8)

        # Treeview (con scrollbar)
        table_frame = tk.Frame(right_panel, bg=Colors.SURFACE)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(6,0))

        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree = ttk.Treeview(table_frame, show="headings", yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(fill=tk.BOTH, expand=True)

        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

    # -----------------------------
    # INTERACCIONES AUXILIARES
    # -----------------------------
    def _handle_back(self):
        """Si existe back_callback lo llamamos; si no, limpiamos el frame."""
        if callable(self.back_callback):
            try:
                self.back_callback()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo volver al inicio: {e}")
        else:
            # destruir los widgets para simular "volver"
            for w in self.parent.winfo_children():
                w.destroy()

    def _on_list_select(self):
        sel = self.report_listbox.curselection()
        if sel:
            idx = sel[0]
            #mapear a selected_report string usado por get_queries()
            # simplemente dejamos seleccionado y al generar reporte usaremos la lista
            # podemos cargar automáticamente al seleccionar:
            self.load_report_data()

    def _apply_quick_search(self):
        """Filtro simple: busca texto en la representación de cada fila (tras generar reporte)."""
        term = self.quick_search.get().strip().lower()
        if not term:
            # si vacío - recargar tabla completa
            self.update_tree(self.current_columns, self.current_rows)
            return

        # Filtrar filas por si alguna celda contiene term
        filtered = []
        for r in self.current_rows:
            # r es dict (resultado fetchall devuelve dicts)
            if any(term in (str(v).lower()) for v in r.values() if v is not None):
                filtered.append(r)
        self.update_tree(self.current_columns, filtered)

    # -----------------------------
    # CARGAR DATOS SEGÚN REPORTE
    # -----------------------------
    def load_report_data(self):
        """Detecta reporte seleccionado en listbox y ejecuta la consulta correspondiente."""
        sel = self.report_listbox.curselection()
        if not sel:
            messagebox.showwarning("Seleccionar reporte", "Seleccione un reporte del panel izquierdo.")
            return
        idx = sel[0]
        report_key = self.reports_list[idx]  # texto, coincide con get_queries keys

        queries = self.get_queries()
        qinfo = queries.get(report_key)
        if not qinfo:
            messagebox.showerror("Error", "No existe consulta para el reporte seleccionado.")
            return

        sql = qinfo["query"]
        columns = qinfo["columns"]

        try:
            rows = self.db.fetchall(sql)
        except Exception as e:
            messagebox.showerror("Error BD", f"No se pudo ejecutar la consulta:\n{e}")
            return

        self.current_columns = columns
        self.current_rows = rows

        self.update_tree(columns, rows)

    # -----------------------------
    # ACTUALIZAR TREEVIEW
    # -----------------------------
    def update_tree(self, columns, rows):
        # limpiar
        for r in self.tree.get_children():
            self.tree.delete(r)

        # configurar columnas
        self.tree["columns"] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor=tk.W)

        # insertar filas
        for r in rows:
            # convertir dict a valores en orden
            vals = tuple(r.get(k) if k in r else next(iter(r.values())) for k in r.keys()) if isinstance(r, dict) else tuple(r)
            # mejor: usar r.values() asumiendo que fetchall retorna RealDictCursor -> dict with keys in query order
            try:
                vals = tuple(r.values())
            except Exception:
                vals = tuple(r)
            self.tree.insert("", tk.END, values=vals)

    # -----------------------------
    # DESCARGAR CSV
    # -----------------------------
    def download_report(self):
        rows = [self.tree.item(r)["values"] for r in self.tree.get_children()]
        if not rows:
            messagebox.showwarning("Descarga", "No hay datos para exportar.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(self.current_columns)
                for row in rows:
                    writer.writerow(row)
            messagebox.showinfo("Descarga", f"Reporte guardado en:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar CSV:\n{e}")

    # -----------------------------
    # EXPORTAR EXCEL (openpyxl)
    # -----------------------------
    def export_to_excel(self):
        if not _HAS_OPENPYXL:
            messagebox.showwarning("openpyxl faltante", "Instale openpyxl: pip install openpyxl")
            return

        if not self.current_rows:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")
            return

        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"

            # encabezados
            for i, h in enumerate(self.current_columns, start=1):
                ws.cell(row=1, column=i, value=h)

            for r_idx, row in enumerate(self.current_rows, start=2):
                # si row es dict
                if isinstance(row, dict):
                    values = list(row.values())
                else:
                    values = list(row)
                for c_idx, val in enumerate(values, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

            wb.save(path)
            messagebox.showinfo("Éxito", f"Reporte exportado a Excel:\n{path}")

        except Exception as e:
            messagebox.showerror("Error Excel", f"No se pudo exportar: {e}")

    # -----------------------------
    # EXPORTAR PDF (usando matplotlib -> figura con tabla)
    # -----------------------------
    def export_to_pdf(self):
        if not self.current_rows:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")
            return

        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not path:
            return

        try:
            # crear figura y tabla
            fig = Figure(figsize=(11, 8.5))
            ax = fig.add_subplot(111)
            ax.axis('off')

            # preparar datos
            headers = self.current_columns
            data = []
            for row in self.current_rows:
                if isinstance(row, dict):
                    data.append([str(v) if v is not None else "" for v in row.values()])
                else:
                    data.append([str(v) if v is not None else "" for v in row])

            # dibujar tabla
            table = ax.table(cellText=data, colLabels=headers, loc='center', cellLoc='left')
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1, 1.2)

            fig.savefig(path, bbox_inches='tight')
            messagebox.showinfo("Éxito", f"PDF generado:\n{path}")
        except Exception as e:
            messagebox.showerror("Error PDF", f"No se pudo exportar a PDF:\n{e}")

    # -----------------------------
    # VENTANA DE GRÁFICOS (matplotlib embebido)
    # -----------------------------
    def show_graph_window(self):
        if not self.current_rows:
            messagebox.showwarning("Gráfico", "Genere primero un reporte para ver el gráfico.")
            return

        # Detectar reporte actual por índice
        sel = self.report_listbox.curselection()
        idx = sel[0] if sel else 2
        key = self.reports_list[idx]

        # Crear ventana
        win = tk.Toplevel(self.parent)
        win.title(f"Gráfico - {key}")
        win.geometry("800x520")

        fig = Figure(figsize=(8, 5), dpi=100)
        ax = fig.add_subplot(111)

        try:
            # Gráficos por tipo de reporte (ejemplos)
            if idx == 2:  # Crecimiento anual (reporte 3)
                # esperamos filas con keys 'anio' y 'nuevos_clientes' o similares
                years = []
                counts = []
                for r in self.current_rows:
                    # r puede ser dict
                    vals = list(r.values()) if isinstance(r, dict) else list(r)
                    # intentar extraer dos primeros campos numéricos
                    years.append(int(float(vals[0])))
                    counts.append(int(vals[1]))
                ax.bar(years, counts)
                ax.set_xlabel("Año")
                ax.set_ylabel("Nuevos clientes")
                ax.set_title("Crecimiento anual de clientes")

            elif idx == 3:  # Preferencias minoristas (reporte 4) -> pie
                labels = []
                sizes = []
                for r in self.current_rows:
                    vals = list(r.values()) if isinstance(r, dict) else list(r)
                    labels.append(str(vals[0]) if vals[0] else "Sin preferencia")
                    sizes.append(int(vals[1]))
                ax.pie(sizes, labels=labels, autopct="%1.1f%%")
                ax.set_title("Preferencias - Clientes Minoristas")

            elif idx == 6:  # Distribución de correos (reporte 7)
                labels = []
                sizes = []
                for r in self.current_rows:
                    vals = list(r.values()) if isinstance(r, dict) else list(r)
                    labels.append(str(vals[0]))
                    sizes.append(int(vals[1]))
                ax.bar(labels, sizes)
                ax.set_title("Distribución de correos por tipo de cliente")
                ax.set_ylabel("Cantidad de correos")
                ax.tick_params(axis='x', rotation=30)

            else:
                # Gráfico genérico: mostrar top N primera columna numérica si existe
                # intentar tomar la segunda columna como valores
                labels = []
                values = []
                for r in self.current_rows[:20]:
                    vals = list(r.values()) if isinstance(r, dict) else list(r)
                    labels.append(str(vals[0])[:20])
                    # buscar primer valor numérico en fila
                    num = None
                    for v in vals[1:]:
                        try:
                            num = float(v)
                            break
                        except Exception:
                            continue
                    values.append(num if num is not None else 0)
                ax.bar(labels, values)
                ax.set_title("Gráfico (genérico)")
                ax.set_ylabel("Valor")
                ax.tick_params(axis='x', rotation=30)

        except Exception as e:
            ax.text(0.5, 0.5, f"Error creando gráfico:\n{e}", ha='center', va='center')
            ax.axis('off')

        # Embeder figura en Tk
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # -----------------------------
    # Diccionario de consultas (idéntico a tu SQL original)
    # -----------------------------
    def get_queries(self):
        return {
            "1. Clientes nuevos por fechas festivas": {
                "query": """
                    SELECT c.Id_contrato, c.Descripcion, c.Fecha_inicio,
                           CASE WHEN EXTRACT(MONTH FROM c.Fecha_inicio) IN (12, 1)
                                THEN 'Campaña Festiva' ELSE 'Otro Periodo' END AS periodo,
                           cc.Razon_Social AS cliente_corporativo
                    FROM Contrato c
                    LEFT JOIN ClienteCorporativo cc ON cc.RUC = c.RUC_Corporativo
                    WHERE EXTRACT(MONTH FROM c.Fecha_inicio) IN (12, 1);
                """,
                "columns": ["Id Contrato", "Descripción", "Fecha Inicio", "Periodo", "Cliente Corporativo"]
            },

            "2. Ranking de regiones activas": {
                "query": """
                    SELECT tipo_cliente,
                           SPLIT_PART(Direccion_Fiscal, '-', 2) AS ciudad,
                           COUNT(*) AS cantidad_clientes
                    FROM (
                        SELECT 'Corporativo' AS tipo_cliente, Direccion_Fiscal FROM DatosClienteCorporativo
                        UNION ALL
                        SELECT 'Mayorista', Direccion_Fiscal FROM ClienteMayorista
                    ) AS direcciones
                    GROUP BY tipo_cliente, ciudad
                    ORDER BY cantidad_clientes DESC;
                """,
                "columns": ["Tipo Cliente", "Ciudad", "Cantidad Clientes"]
            },

            "3. Crecimiento anual de clientes": {
                "query": """
                    SELECT EXTRACT(YEAR FROM Fecha_inicio) AS anio,
                           COUNT(Id_contrato) AS nuevos_clientes
                    FROM Contrato
                    GROUP BY anio
                    ORDER BY anio;
                """,
                "columns": ["Año", "Nuevos Clientes"]
            },

            "4. Preferencias de clientes minoristas": {
                "query": """
                    SELECT Preferencias, COUNT(*) AS cantidad
                    FROM ClienteMinorista
                    GROUP BY Preferencias
                    ORDER BY cantidad DESC;
                """,
                "columns": ["Preferencias", "Cantidad"]
            },

            "5. Clientes sin contrato completo": {
                "query": """
                    SELECT 
                        'Corporativo' AS tipo_cliente,
                        cc.Razon_Social AS nombre_cliente,
                        CASE WHEN c.Id_contrato IS NULL THEN 'Sin Contrato' ELSE 'Con Contrato' END AS estado_contrato,
                        CASE WHEN dcc.Direccion_Fiscal IS NULL OR dcc.Telefono IS NULL THEN 'Datos incompletos' ELSE 'Datos completos' END AS estado_datos
                    FROM ClienteCorporativo cc
                    LEFT JOIN Contrato c ON cc.RUC = c.RUC_Corporativo
                    LEFT JOIN DatosClienteCorporativo dcc ON cc.RUC = dcc.RUC_Corporativo

                    UNION ALL

                    SELECT
                        'Mayorista' AS tipo_cliente,
                        cm.Razon_Social AS nombre_cliente,
                        'Sin Contrato' AS estado_contrato,
                        CASE WHEN dcm.Telefono IS NULL OR dcm.Correo IS NULL THEN 'Datos incompletos' ELSE 'Datos completos' END AS estado_datos
                    FROM ClienteMayorista cm
                    LEFT JOIN DatosClienteMayorista dcm ON cm.RUC = dcm.RUC_Mayorista;
                """,
                "columns": ["Tipo Cliente", "Nombre", "Estado Contrato", "Estado Datos"]
            },

            "6. Cobertura de clientes corporativos y mayoristas": {
                "query": """
                    SELECT tipo_cliente, nombre_cliente,
                           COUNT(DISTINCT TRIM(ciudad)) AS cantidad_ciudades
                    FROM (
                        SELECT 'Mayorista' AS tipo_cliente, cm.Razon_Social AS nombre_cliente, SPLIT_PART(cm.Direccion_Fiscal, '-', 2) AS ciudad FROM ClienteMayorista cm
                        UNION ALL
                        SELECT 'Corporativo', cc.Razon_Social, SPLIT_PART(dcc.Direccion_Fiscal, '-', 2) FROM ClienteCorporativo cc JOIN DatosClienteCorporativo dcc ON cc.RUC = dcc.RUC_Corporativo
                    ) AS cobertura
                    GROUP BY tipo_cliente, nombre_cliente
                    ORDER BY cantidad_ciudades DESC;
                """,
                "columns": ["Tipo Cliente", "Cliente", "Ciudades Cubiertas"]
            },

            "7. Distribución de correos electrónicos": {
                "query": """
                    SELECT 'Cliente Minorista' AS tipo, COUNT(Correo) AS total FROM ClienteMinorista
                    UNION ALL
                    SELECT 'Cliente Mayorista', COUNT(Correo) FROM DatosClienteMayorista
                    UNION ALL
                    SELECT 'Cliente Corporativo', COUNT(Correo) FROM ClienteCorporativo;
                """,
                "columns": ["Tipo Cliente", "Cantidad Correos"]
            },

            "8. Contratos activos por cliente corporativo": {
                "query": """
                    SELECT cc.Razon_Social AS cliente, COUNT(c.Id_contrato) AS total_activos
                    FROM ClienteCorporativo cc
                    LEFT JOIN Contrato c ON cc.RUC = c.RUC_Corporativo
                    WHERE c.Estado = 'Activo'
                    GROUP BY cc.Razon_Social
                    ORDER BY total_activos DESC
                    LIMIT 10;
                """,
                "columns": ["Cliente Corporativo", "Contratos Activos"]
            },

            "9. Administradores con más clientes corporativos activos": {
                "query": """
                    SELECT a.DNI, a.Nombre_Apellido AS administrador, COUNT(DISTINCT cc.RUC) AS clientes_activos
                    FROM Administrador a
                    LEFT JOIN ClienteCorporativo cc ON a.DNI = cc.DNI_administrador
                    LEFT JOIN Contrato c ON cc.RUC = c.RUC_Corporativo
                    WHERE c.Estado = 'Activo'
                    GROUP BY a.DNI, a.Nombre_Apellido
                    ORDER BY clientes_activos DESC;
                """,
                "columns": ["DNI", "Administrador", "Clientes Activos"]
            },

            "10. Ranking de administradores según clientes gestionados": {
                "query": """
                    SELECT a.DNI, a.Nombre_Apellido AS administrador, COUNT(todos.DNI) AS cantidad
                    FROM Administrador a
                    LEFT JOIN (
                        SELECT DNI_administrador AS DNI FROM ClienteMinorista
                        UNION ALL
                        SELECT DNI_administrador FROM ClienteMayorista
                        UNION ALL
                        SELECT DNI_administrador FROM ClienteCorporativo
                    ) todos ON a.DNI = todos.DNI
                    GROUP BY a.DNI, a.Nombre_Apellido
                    ORDER BY cantidad DESC;
                """,
                "columns": ["DNI", "Administrador", "Clientes Gestionados"]
            }
        }
